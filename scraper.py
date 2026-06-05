import re
import json
import time
import html as ihtml
import logging
from urllib.parse import urljoin, urlparse, parse_qsl, urlencode, urlunparse
from xml.etree import ElementTree as ET

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL = "https://apisexpress.com"
SITEMAP_INDEX = f"{BASE_URL}/sitemap.xml"
OUTFILE = "apisexpress_all_products.xlsx"

REQUEST_TIMEOUT = 30
SLEEP_SECONDS = 0.15
MAX_RETRIES = 3

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; APISExpressScraper/2.0; +https://github.com/)",
    "Accept-Language": "bg-BG,bg;q=0.9,en;q=0.8",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")

session = requests.Session()
session.headers.update(HEADERS)


def add_nocache(url: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query))
    query["_nocache"] = str(int(time.time()))
    return urlunparse(parsed._replace(query=urlencode(query)))


def fetch(url: str, nocache: bool = False) -> str:
    if nocache:
        url = add_nocache(url)

    last_error = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            return r.text
        except Exception as e:
            last_error = e
            logging.warning("Fetch failed (%s/%s): %s -> %s", attempt, MAX_RETRIES, url, e)
            time.sleep(1.5 * attempt)

    raise RuntimeError(f"Could not fetch {url}: {last_error}")


def parse_xml_locs(xml_text: str) -> list[str]:
    root = ET.fromstring(xml_text.encode("utf-8"))
    locs = []

    for elem in root.iter():
        if elem.tag.endswith("loc") and elem.text:
            locs.append(elem.text.strip())

    return locs


def get_product_urls() -> list[str]:
    logging.info("Reading sitemap index: %s", SITEMAP_INDEX)

    index_xml = fetch(SITEMAP_INDEX)
    sitemap_locs = parse_xml_locs(index_xml)

    product_sitemaps = [
        u for u in sitemap_locs
        if re.search(r"/product-sitemap\d*\.xml$", u)
    ]

    logging.info("Found %s product sitemap(s): %s", len(product_sitemaps), ", ".join(product_sitemaps))

    product_urls = []

    for sm in product_sitemaps:
        xml = fetch(sm)
        urls = [u for u in parse_xml_locs(xml) if "/produkt/" in u]
        logging.info("%s -> %s product URLs", sm, len(urls))
        product_urls.extend(urls)
        time.sleep(SLEEP_SECONDS)

    seen = set()
    unique = []

    for u in product_urls:
        if u not in seen:
            seen.add(u)
            unique.append(u)

    logging.info("Total unique product URLs: %s", len(unique))
    return unique


def clean_text(value) -> str:
    if value is None:
        return ""

    if isinstance(value, list):
        value = " ".join(str(x) for x in value)

    text = BeautifulSoup(str(value), "html.parser").get_text(" ", strip=True)
    return re.sub(r"\s+", " ", ihtml.unescape(text)).strip()


def clean_eur_price(text: str) -> str:
    if not text:
        return ""

    text = ihtml.unescape(text)
    text = text.replace("\xa0", " ")
    text = text.replace("€", " ")
    text = text.replace("лв.", " ")
    text = text.replace("лв", " ")
    text = text.replace("(", " ")
    text = text.replace(")", " ")
    text = re.sub(r"\s+", " ", text).strip()

    match = re.search(r"\d+(?:[.,]\d+)?", text)
    return match.group(0).replace(",", ".") if match else ""


def extract_price_eur(soup: BeautifulSoup, product_jsonld: dict) -> str:
    """
    Взима цената само от основния продукт.
    Ако има намаление, взима намалената цена от <ins>.
    """

    main_product = soup.select_one("div[id^='product-'].single-product-page")
    if not main_product:
        main_product = soup.select_one("div[id^='product-']")

    if not main_product:
        return ""

    price_box = main_product.select_one(
        ".summary-inner p.price, "
        ".summary p.price, "
        ".entry-summary p.price, "
        "p.price"
    )

    if not price_box:
        return ""

    sale_price = price_box.select_one("ins .woocommerce-Price-amount")

    if sale_price:
        sale_text = sale_price.get_text(" ", strip=True)
        if "€" in sale_text:
            return clean_eur_price(sale_text)

    amounts = price_box.select(".woocommerce-Price-amount")

    for amount in amounts:
        amount_text = amount.get_text(" ", strip=True)

        if "€" not in amount_text:
            continue

        if amount.find_parent("del"):
            continue

        if "лв" in amount_text.lower():
            continue

        return clean_eur_price(amount_text)

    return ""


def uniq(seq):
    out = []
    seen = set()

    for x in seq:
        if not x:
            continue

        x = x.strip()

        if x and x not in seen:
            seen.add(x)
            out.append(x)

    return out


def normalize_image_url(url: str) -> str:
    if not url:
        return ""

    url = ihtml.unescape(url).strip().split()[0]
    url = urljoin(BASE_URL, url)

    # Връща оригиналния upload вместо WordPress thumbnail размери
    url = re.sub(r"-\d+x\d+(?=\.(?:jpg|jpeg|png|webp|gif)$)", "", url, flags=re.I)

    return url


def find_jsonld_objects(soup: BeautifulSoup):
    objects = []

    for script in soup.find_all("script", attrs={"type": "application/ld+json"}):
        raw = script.string or script.get_text("", strip=True)

        if not raw:
            continue

        try:
            data = json.loads(ihtml.unescape(raw))
        except Exception:
            continue

        stack = [data]

        while stack:
            item = stack.pop()

            if isinstance(item, dict):
                objects.append(item)

                for v in item.values():
                    if isinstance(v, (dict, list)):
                        stack.append(v)

            elif isinstance(item, list):
                stack.extend(item)

    return objects


def type_matches(obj: dict, wanted: str) -> bool:
    t = obj.get("@type")

    if isinstance(t, list):
        return wanted in t

    return t == wanted


def get_product_jsonld(objects):
    for obj in objects:
        if type_matches(obj, "Product"):
            return obj

    return {}


def get_breadcrumb_categories(objects, product_name: str) -> str:
    cats = []

    for obj in objects:
        if not type_matches(obj, "BreadcrumbList"):
            continue

        elements = obj.get("itemListElement") or []
        names = []

        for el in elements:
            item = el.get("item") if isinstance(el, dict) else None
            name = ""

            if isinstance(item, dict):
                name = item.get("name", "")
            elif isinstance(el, dict):
                name = el.get("name", "")

            name = clean_text(name)

            if name:
                names.append(name)

        for name in names:
            low = name.lower()

            if low in {"home", "начало", "магазин"}:
                continue

            if product_name and clean_text(product_name) and name == clean_text(product_name):
                continue

            if name not in cats:
                cats.append(name)

    return " | ".join(cats)


def availability_bg(value: str, soup: BeautifulSoup) -> str:
    text = (value or "").lower()
    page_text = soup.get_text(" ", strip=True).lower()

    product_div = soup.select_one("div[id^='product-']")
    classes = " ".join(product_div.get("class", [])) if product_div else ""

    if "outofstock" in text or "out-of-stock" in text or "out of stock" in text:
        return "Не е наличен"

    if "instock" in text or "in stock" in text:
        return "Наличен"

    if (
        "в момента този артикул не е наличен" in page_text
        or "изчерпан" in page_text
        or "outofstock" in classes
        or "out-of-stock" in classes
    ):
        return "Не е наличен"

    if "instock" in classes or "in-stock" in classes:
        return "Наличен"

    return ""


def parse_product(url: str) -> dict:
    html = fetch(url, nocache=True)
    soup = BeautifulSoup(html, "html.parser")

    objects = find_jsonld_objects(soup)
    product = get_product_jsonld(objects)

    name = clean_text(product.get("name"))

    if not name:
        h1 = soup.select_one("h1.product_title, h1.entry-title, h1")
        name = clean_text(h1.get_text(" ", strip=True) if h1 else "")

    sku = clean_text(product.get("sku"))

    if not sku:
        sku_el = soup.select_one(".sku")
        sku = clean_text(sku_el.get_text(" ", strip=True) if sku_el else "")

    if not sku:
        m = re.search(r'data-product_sku="([^"]*)"', html)
        sku = clean_text(m.group(1)) if m else ""

    description = clean_text(product.get("description"))

    if not description:
        desc_el = soup.select_one(
            "#tab-description, "
            ".woocommerce-product-details__short-description, "
            ".product-short-description"
        )
        description = clean_text(desc_el.get_text(" ", strip=True) if desc_el else "")

    if not description:
        meta_desc = soup.select_one('meta[name="description"]')
        description = clean_text(meta_desc.get("content", "") if meta_desc else "")

    price = extract_price_eur(soup, product)

    offers = product.get("offers") or {}

    if isinstance(offers, list):
        offers = offers[0] if offers else {}

    availability = availability_bg(str(offers.get("availability", "")), soup)

    categories = get_breadcrumb_categories(objects, name)

    if not categories:
        product_div = soup.select_one("div[id^='product-']")

        if product_div:
            slugs = [
                c.replace("product_cat-", "")
                for c in product_div.get("class", [])
                if c.startswith("product_cat-")
            ]
            categories = " | ".join(slugs)

    images = []

    img_data = product.get("image")

    if isinstance(img_data, str):
        images.append(normalize_image_url(img_data))

    elif isinstance(img_data, list):
        for im in img_data:
            if isinstance(im, str):
                images.append(normalize_image_url(im))
            elif isinstance(im, dict):
                images.append(normalize_image_url(im.get("url") or im.get("contentUrl") or ""))

    elif isinstance(img_data, dict):
        images.append(normalize_image_url(img_data.get("url") or img_data.get("contentUrl") or ""))

    for meta in soup.select('meta[property="og:image"], meta[property="og:image:secure_url"]'):
        images.append(normalize_image_url(meta.get("content", "")))

    for a in soup.select(".woocommerce-product-gallery a[href], .product-images a[href]"):
        href = a.get("href", "")

        if re.search(r"\.(jpg|jpeg|png|webp|gif)(\?|$)", href, re.I):
            images.append(normalize_image_url(href))

    for img in soup.select(".woocommerce-product-gallery img, .product-images img, img.wp-post-image"):
        for attr in ["data-large_image", "data-src", "src"]:
            images.append(normalize_image_url(img.get(attr, "")))

        srcset = img.get("srcset", "")

        for candidate in srcset.split(","):
            candidate = candidate.strip()

            if candidate:
                images.append(normalize_image_url(candidate.split(" ")[0]))

    images = [u for u in uniq(images) if "/wp-content/uploads/" in u]

    row = {
        "Код на продукт": sku,
        "Име на продукт": name,
        "Категория": categories,
        "Цена в евро": price,
        "Наличност": availability,
        "Описание": description,
        "URL": url,
    }

    for idx, image_url in enumerate(images, start=1):
        row[f"Изображение {idx}"] = image_url

    return row


def save_xlsx(rows: list[dict], filename: str):
    base_cols = [
        "Код на продукт",
        "Име на продукт",
        "Категория",
        "Цена в евро",
        "Наличност",
        "Описание",
        "URL",
    ]

    max_img = 0

    for r in rows:
        for k in r.keys():
            m = re.match(r"Изображение (\d+)$", k)

            if m:
                max_img = max(max_img, int(m.group(1)))

    cols = base_cols + [f"Изображение {i}" for i in range(1, max_img + 1)]

    df = pd.DataFrame(rows)

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    df = df[cols]
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active
    ws.title = "Products"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 18,
        "B": 55,
        "C": 45,
        "D": 12,
        "E": 16,
        "F": 70,
        "G": 55,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for col_idx in range(8, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 60

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(filename)


def main():
    product_urls = get_product_urls()

    rows = []
    errors = []

    for i, url in enumerate(product_urls, start=1):
        logging.info("[%s/%s] %s", i, len(product_urls), url)

        try:
            rows.append(parse_product(url))
        except Exception as e:
            logging.exception("Failed product: %s", url)
            errors.append({"URL": url, "Грешка": str(e)})

        time.sleep(SLEEP_SECONDS)

    save_xlsx(rows, OUTFILE)

    logging.info("Saved %s with %s products. Errors: %s", OUTFILE, len(rows), len(errors))

    if errors:
        pd.DataFrame(errors).to_excel("apisexpress_errors.xlsx", index=False)
        logging.info("Saved apisexpress_errors.xlsx")


if __name__ == "__main__":
    main()
